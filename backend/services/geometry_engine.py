import pulp
import numpy as np
import pandas as pd
import io
import datetime
import matplotlib
# Configurar backend 'Agg' para que funcione en servidor sin interfaz gráfica
matplotlib.use('Agg') 
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

class LinearProgrammingEngine:
    
    @staticmethod
    def solve_and_generate_geometry(objective, constraints, maximize=True):
        """
        Calcula la solución óptima y genera los datos geométricos para el gráfico.
        """
        # 1. RESOLVER MATEMÁTICAMENTE CON PULP
        prob_type = pulp.LpMaximize if maximize else pulp.LpMinimize
        prob = pulp.LpProblem("Graphic_LP", prob_type)
        
        x = pulp.LpVariable("x", lowBound=0)
        y = pulp.LpVariable("y", lowBound=0)
        
        # Función Objetivo
        prob += objective['x'] * x + objective['y'] * y, "Objective_Function"
        
        # Restricciones
        for i, c in enumerate(constraints):
            lhs = c['x'] * x + c['y'] * y
            rhs = c['rhs']
            if c['operator'] == '<=': prob += lhs <= rhs, f"C{i}"
            elif c['operator'] == '>=': prob += lhs >= rhs, f"C{i}"
            elif c['operator'] == '=': prob += lhs == rhs, f"C{i}"

        # Solver
        prob.solve(pulp.PULP_CBC_CMD(msg=False))
        status = pulp.LpStatus[prob.status]
        
        optimal_point = None
        optimal_z = 0
        
        if status == 'Optimal':
            val_x = x.varValue if x.varValue is not None else 0.0
            val_y = y.varValue if y.varValue is not None else 0.0
            optimal_point = {'x': val_x, 'y': val_y}
            optimal_z = pulp.value(prob.objective)

        # 2. CÁLCULO DE GEOMETRÍA (Intersecciones)
        lines = []
        lines.append({'x': 1, 'y': 0, 'rhs': 0, 'type': 'axis'}) # x=0
        lines.append({'x': 0, 'y': 1, 'rhs': 0, 'type': 'axis'}) # y=0
        
        for c in constraints:
            lines.append(c)

        points = []
        
        # Encontrar intersecciones
        for i in range(len(lines)):
            for j in range(i + 1, len(lines)):
                l1 = lines[i]
                l2 = lines[j]
                
                try:
                    A = np.array([[l1['x'], l1['y']], [l2['x'], l2['y']]])
                    B = np.array([l1['rhs'], l2['rhs']])
                    
                    intersection = np.linalg.solve(A, B)
                    pt_x, pt_y = intersection[0], intersection[1]
                    
                    if LinearProgrammingEngine._is_feasible(pt_x, pt_y, constraints):
                        is_duplicate = False
                        for p in points:
                            if np.isclose(p['x'], pt_x, atol=1e-4) and np.isclose(p['y'], pt_y, atol=1e-4):
                                is_duplicate = True
                                break
                        if not is_duplicate:
                            points.append({'x': float(pt_x), 'y': float(pt_y)})
                            
                except np.linalg.LinAlgError:
                    continue 

        if len(points) > 2:
            points = LinearProgrammingEngine._sort_clockwise(points)

        return {
            "status": status,
            "optimal_solution": {
                "point": optimal_point,
                "z_value": optimal_z
            },
            "feasible_region": points,
            "constraints_lines": constraints
        }

    @staticmethod
    def _is_feasible(x, y, constraints):
        tol = 1e-5
        if x < -tol or y < -tol: return False
            
        for c in constraints:
            val = c['x'] * x + c['y'] * y
            if c['operator'] == '<=':
                if val > c['rhs'] + tol: return False
            elif c['operator'] == '>=':
                if val < c['rhs'] - tol: return False
            elif c['operator'] == '=':
                if not np.isclose(val, c['rhs'], atol=tol): return False
        return True

    @staticmethod
    def _sort_clockwise(points):
        center_x = sum(p['x'] for p in points) / len(points)
        center_y = sum(p['y'] for p in points) / len(points)
        def sort_key(p): return np.arctan2(p['y'] - center_y, p['x'] - center_x)
        return sorted(points, key=sort_key)

    @staticmethod
    def generate_excel_report(result_data, objective, constraints, maximize):
        """
        Genera un reporte Excel completo con 4 HOJAS: Resumen, Análisis, Gráfico Zoom, Gráfico Completo.
        """
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        # PREPARACIÓN DE DATOS
        opt_point = result_data['optimal_solution'].get('point')
        opt_x = opt_point['x'] if opt_point else 0
        opt_y = opt_point['y'] if opt_point else 0
        opt_z = result_data['optimal_solution'].get('z_value', 0)
        
        # --- HOJA 1: RESUMEN ---
        data_resumen = [
            ['Estado', result_data['status']],
            ['Objetivo', 'Maximizar' if maximize else 'Minimizar'],
            ['Función Z', f"Z = {objective['x']}x1 + {objective['y']}x2"],
            ['Valor Óptimo (Z)', opt_z],
            ['Valor X1', opt_x],
            ['Valor X2', opt_y],
            ['Fecha Generación', datetime.datetime.now().strftime("%Y-%m-%d %H:%M")]
        ]
        df_resumen = pd.DataFrame(data_resumen, columns=['Concepto', 'Resultado'])
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        # --- HOJA 2: ANÁLISIS ---
        analysis_data = []
        for i, c in enumerate(constraints):
            uso = c['x'] * opt_x + c['y'] * opt_y
            limite = c['rhs']
            holgura = abs(limite - uso)
            es_activa = holgura < 1e-5
            estado_texto = "RECURSO AGOTADO (Limitante)" if es_activa else "RECURSO DISPONIBLE (Holgura)"
            if c['operator'] == '=': estado_texto = "CUMPLIDA (Igualdad)"
            
            analysis_data.append({
                'ID': f"R{i+1}",
                'Ecuación': f"{c['x']}x1 + {c['y']}x2 {c['operator']} {c['rhs']}",
                'Valor Lado Izq (Uso)': round(uso, 4),
                'Valor Lado Der (Límite)': limite,
                'Holgura / Excedente': round(holgura, 4),
                'Estado': estado_texto
            })
            
        df_analysis = pd.DataFrame(analysis_data)
        df_analysis.to_excel(writer, sheet_name='Análisis', index=False)

        # --- HOJA 3: GRÁFICO ZOOM (Región Factible) ---
        wb = writer.book
        ws_graph1 = wb.create_sheet("Gráfico Zoom (Factible)")
        
        img_buffer1 = LinearProgrammingEngine._create_matplotlib_plot(
            result_data['feasible_region'], constraints, objective, opt_point, maximize, mode='focused'
        )
        
        if img_buffer1:
            img1 = OpenpyxlImage(img_buffer1)
            ws_graph1.add_image(img1, 'B2')
            ws_graph1['A1'] = "Vista Detallada: Región Factible y Solución"
        else:
            ws_graph1['A1'] = "No se pudo generar el gráfico."

        # --- HOJA 4: GRÁFICO COMPLETO (Mapa Global) ---
        ws_graph2 = wb.create_sheet("Gráfico Completo (Mapa)")
        
        img_buffer2 = LinearProgrammingEngine._create_matplotlib_plot(
            result_data['feasible_region'], constraints, objective, opt_point, maximize, mode='full'
        )
        
        if img_buffer2:
            img2 = OpenpyxlImage(img_buffer2)
            ws_graph2.add_image(img2, 'B2')
            ws_graph2['A1'] = "Vista General: Mapa de Restricciones Completo"
        else:
            ws_graph2['A1'] = "No se pudo generar el gráfico."

        # --- FORMATO FINAL (Autoajuste) ---
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        writer.close()
        output.seek(0)
        return output

    @staticmethod
    def _create_matplotlib_plot(feasible_region, constraints, objective, opt_point, maximize, mode='focused'):
        """
        Crea un plot de Matplotlib.
        mode='focused': Zoom ajustado a la región factible y la solución.
        mode='full': Zoom alejado para ver los cortes con los ejes de todas las restricciones.
        """
        try:
            fig, ax = plt.subplots(figsize=(10, 8))
            
            # --- LÓGICA DE LÍMITES DINÁMICOS ---
            limit_x, limit_y = 10, 10 # Defaults
            
            if mode == 'focused':
                # Modo original: basado solo en los puntos visibles de la solución
                all_x = [p['x'] for p in feasible_region] if feasible_region else [0]
                all_y = [p['y'] for p in feasible_region] if feasible_region else [0]
                if opt_point:
                    all_x.append(opt_point['x'])
                    all_y.append(opt_point['y'])
                
                max_x = max(all_x) if all_x else 5
                max_y = max(all_y) if all_y else 5
                
                limit_x = max_x * 1.25 if max_x > 0 else 10
                limit_y = max_y * 1.25 if max_y > 0 else 10
                
            else: # mode == 'full'
                # Modo completo: Buscar interceptos con los ejes de TODAS las restricciones
                max_intercept_x = 0
                max_intercept_y = 0
                
                # Revisamos puntos factibles también para asegurar que estén dentro
                if feasible_region:
                    max_intercept_x = max(p['x'] for p in feasible_region)
                    max_intercept_y = max(p['y'] for p in feasible_region)
                
                for c in constraints:
                    # Intercepto en X (cuando y=0): ax <= rhs => x = rhs/a
                    if abs(c['x']) > 1e-5: 
                        val = c['rhs'] / c['x']
                        if val > max_intercept_x: max_intercept_x = val
                    
                    # Intercepto en Y (cuando x=0): by <= rhs => y = rhs/b
                    if abs(c['y']) > 1e-5:
                        val = c['rhs'] / c['y']
                        if val > max_intercept_y: max_intercept_y = val
                
                # Si los interceptos son muy bestias (ej: x <= 10000), limitamos para no romper el gráfico
                # Pero en Simplex académico se suele querer ver todo.
                limit_x = max_intercept_x * 1.1 if max_intercept_x > 0 else 10
                limit_y = max_intercept_y * 1.1 if max_intercept_y > 0 else 10
            
            # Aplicar límites
            ax.set_xlim(-1, limit_x)
            ax.set_ylim(-1, limit_y)
            
            # --- DIBUJAR ---
            x_vals = np.linspace(-1, limit_x, 200)
            colors = ['blue', 'green', 'orange', 'purple', 'brown', 'teal', 'magenta']
            
            # 1. Restricciones
            for i, c in enumerate(constraints):
                color = colors[i % len(colors)]
                label = f"R{i+1}: {c['x']}x₁ + {c['y']}x₂ {c['operator']} {c['rhs']}"
                
                if abs(c['y']) > 1e-5:
                    y_vals = (c['rhs'] - c['x'] * x_vals) / c['y']
                    # Filtrar valores fuera de rango visual para limpieza
                    valid_mask = (y_vals >= -limit_y) & (y_vals <= limit_y * 1.5)
                    ax.plot(x_vals[valid_mask], y_vals[valid_mask], label=label, color=color, alpha=0.6, linewidth=2)
                else:
                    val_x = c['rhs'] / c['x']
                    ax.axvline(x=val_x, label=label, color=color, alpha=0.6, linewidth=2)

            # 2. Región Factible
            if feasible_region and len(feasible_region) > 2:
                poly_x = [p['x'] for p in feasible_region]
                poly_y = [p['y'] for p in feasible_region]
                poly_x.append(poly_x[0])
                poly_y.append(poly_y[0])
                ax.fill(poly_x, poly_y, 'gray', alpha=0.3, label='Región Factible')
                
            # 3. Solución Óptima
            if opt_point:
                ax.plot(opt_point['x'], opt_point['y'], 'ro', markersize=12, label='Solución Óptima', zorder=5, markeredgecolor='white', markeredgewidth=2)
                
                # Etiqueta mejorada
                annotation_text = f"Óptimo\n({opt_point['x']:.1f}, {opt_point['y']:.1f})"
                ax.annotate(annotation_text, 
                            (opt_point['x'], opt_point['y']),
                            xytext=(15, 15), textcoords='offset points',
                            bbox=dict(boxstyle="round,pad=0.5", fc="#FFEB3B", alpha=0.9, ec="orange"),
                            arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=.2"))

            # Títulos y estilos
            title_suffix = "(Vista Detallada)" if mode == 'focused' else "(Mapa Completo)"
            ax.set_title(f"Gráfico Método Simplex {title_suffix}", fontsize=16, pad=15)
            ax.set_xlabel("Variable X₁", fontsize=12)
            ax.set_ylabel("Variable X₂", fontsize=12)
            
            # Grid mejorado
            ax.grid(True, which='major', linestyle='-', alpha=0.8, color='#ddd')
            ax.minorticks_on()
            ax.grid(True, which='minor', linestyle=':', alpha=0.4, color='#eee')
            
            # Ejes principales marcados
            ax.axhline(0, color='black', linewidth=1.5)
            ax.axvline(0, color='black', linewidth=1.5)
            
            # Leyenda fuera si hay muchas restricciones
            loc_legend = 'upper right' if mode == 'focused' else 'best'
            ax.legend(loc=loc_legend, fancybox=True, shadow=True, fontsize='small', framealpha=0.9)
            
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', bbox_inches='tight', dpi=120) # Mayor DPI para mejor calidad
            img_buffer.seek(0)
            plt.close(fig)
            return img_buffer
            
        except Exception as e:
            print(f"Error generando plot ({mode}): {e}")
            return None